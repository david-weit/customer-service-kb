"""Excel 解析器（仿 RAGFlow RAGFlowExcelParser）。"""

import logging
import re
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, List, Union

import pandas as pd
from openpyxl import Workbook, load_workbook

ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


class ExcelParser:
    """将表格行转为「列名：值」自然语言，便于检索。"""

    @staticmethod
    def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        def clean_string(s):
            if isinstance(s, str):
                return ILLEGAL_CHARACTERS_RE.sub(" ", s)
            return s

        return df.apply(lambda col: col.map(clean_string))

    @staticmethod
    def _fill_worksheet_from_dataframe(ws, df: pd.DataFrame) -> None:
        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)
        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

    @staticmethod
    def _dataframe_to_workbook(df) -> Workbook:
        if isinstance(df, dict) and len(df) > 1:
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            for sheet_name, frame in df.items():
                frame = ExcelParser._clean_dataframe(frame)
                ws = wb.create_sheet(title=str(sheet_name)[:31])
                ExcelParser._fill_worksheet_from_dataframe(ws, frame)
            return wb

        df = ExcelParser._clean_dataframe(df)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ExcelParser._fill_worksheet_from_dataframe(ws, df)
        return wb

    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)
        if isinstance(file_like_object, (str, Path)):
            file_like_object = open(file_like_object, "rb")
            close_after = True
        else:
            close_after = False

        try:
            file_like_object.seek(0)
            file_head = file_like_object.read(4)
            file_like_object.seek(0)

            if not (
                file_head.startswith(b"PK\x03\x04")
                or file_head.startswith(b"\xd0\xcf\x11\xe0")
            ):
                df = pd.read_csv(file_like_object, on_bad_lines="skip")
                return ExcelParser._dataframe_to_workbook(df)

            try:
                return load_workbook(file_like_object, data_only=True)
            except Exception as e:
                logging.info("openpyxl load error: %s, try pandas instead", e)
                file_like_object.seek(0)
                dfs = pd.read_excel(file_like_object, sheet_name=None)
                return ExcelParser._dataframe_to_workbook(dfs)
        finally:
            if close_after:
                file_like_object.close()

    @staticmethod
    def _get_actual_row_count(ws) -> int:
        max_row = ws.max_row or 0
        if max_row <= 10000:
            return max_row

        max_col = min(ws.max_column or 1, 50)

        def row_has_data(row_idx: int) -> bool:
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    return True
            return False

        last_data_row = 0
        for r in range(1, max_row + 1):
            if row_has_data(r):
                last_data_row = r
        return last_data_row

    def __call__(self, fnm: Union[str, Path, bytes, BinaryIO]) -> List[str]:
        wb = self._load_excel_to_workbook(fnm)
        res: List[str] = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            try:
                actual_rows = self._get_actual_row_count(ws)
                rows = list(ws.iter_rows(min_row=1, max_row=actual_rows)) if actual_rows else []
            except Exception as e:
                logging.warning("Skip sheet '%s' due to rows access error: %s", sheetname, e)
                continue
            if not rows:
                continue

            headers = list(rows[0])
            for row in rows[1:]:
                fields = []
                for i, cell in enumerate(row):
                    if cell.value is None or str(cell.value).strip() == "":
                        continue
                    header = str(headers[i].value) if i < len(headers) and headers[i].value is not None else ""
                    fields.append(f"{header}：{cell.value}" if header else str(cell.value))
                if not fields:
                    continue
                line = "; ".join(fields)
                if "sheet" not in sheetname.lower():
                    line += " ——" + sheetname
                res.append(line)
        return res
